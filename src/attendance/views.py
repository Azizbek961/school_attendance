# attendance/views.py
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
import openpyxl
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.conf import settings



def login_view(request):
    if request.user.is_authenticated:
        # Superuser uchun alohida tekshiruv
        if request.user.is_superuser:
            return redirect('manage_users')
    # Agar foydalanuvchi allaqachon login qilgan bo'lsa, uni dashboardga yo'naltirish
    if request.user.is_authenticated:
        # Foydalanuvchi roliga qarab yo'naltirish
        if hasattr(request.user, 'profile'):
            if request.user.profile.role == 'student':
                return redirect('student_dashboard')
            elif request.user.profile.role == 'teacher':
                return redirect('teacher_dashboard')
            elif request.user.profile.role == 'admin':
                return redirect('manage_users')
        # Agar profile bo'lmasa, oddiy foydalanuvchilar uchun
        pass

    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user_type = request.POST.get('user_type', 'student')
        remember_me = request.POST.get('remember_me') == 'on'

        # Foydalanuvchini autentifikatsiya qilish
        user = authenticate(request, username=username, password=password)

        if user is not None:
            # Login qilish
            login(request, user)

            # "Remember me" ni sozlash
            if not remember_me:
                request.session.set_expiry(0)  # Session brauzer yopilganda tugaydi
            else:
                request.session.set_expiry(1209600)  # 2 hafta

            # Foydalanuvchi roliga qarab yo'naltirish
            if hasattr(user, 'profile'):
                if user.profile.role == 'student':
                    return redirect('student_dashboard')
                elif user.profile.role == 'teacher':
                    return redirect('teacher_dashboard')
                elif user.profile.role == 'admin':
                    return redirect('manage_users')

            # Agar profile bo'lmasa, user_type bo'yicha yo'naltirish
            messages.success(request, f"Xush kelibsiz, {user.username}!")
            if user_type == 'student':
                return redirect('student_dashboard')
            elif user_type == 'teacher':
                return redirect('teacher_dashboard')
            elif user_type == 'admin':
                return redirect('manage_users')
            else:
                pass
        else:
            # Xato xabar
            messages.error(request, 'Login yoki parol noto\'g\'ri. Iltimos, qayta urinib ko\'ring.')

    # Bo'sh form yoki GET so'rovi uchun
    from .forms import LoginForm
    form = LoginForm()

    return render(request, 'attendance/login.html', {
        'form': form,
        'debug': settings.DEBUG  # Demo hisoblarni faqat development rejimida ko'rsatish
    })


# Logout funksiyasi
from django.contrib.auth import logout


def logout_view(request):
    logout(request)
    messages.success(request, 'Siz tizimdan chiqdingiz.')
    return redirect('login')


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib import messages
from django.db.models import Count, Q, Avg
from django.utils import timezone
from datetime import datetime, timedelta
from django.http import JsonResponse, HttpResponse
from .models import Profile, Class, Subject, Attendance, Settings
from .forms import UserForm, ClassForm, SubjectForm, AttendanceForm, SettingsForm
import csv


def user_is_admin(user):
    """Check if user is admin"""
    # Agar superuser bo'lsa, admin deb hisoblash
    if user.is_superuser:
        return True

    # Agar profile bo'lsa va role 'admin' bo'lsa
    if hasattr(user, 'profile'):
        return user.profile.role == 'admin'

    return False


@login_required
def manage_users(request):
    """Admin dashboard view"""
    if not user_is_admin(request.user):
        messages.error(request, "Sizda admin huquqlari yo'q!")
        return redirect('login')

    # Statistics
    total_students = Profile.objects.filter(role='student').count()
    total_teachers = Profile.objects.filter(role='teacher').count()
    total_classes = Class.objects.count()

    # Today's attendance
    today = timezone.now().date()
    today_attendance = Attendance.objects.filter(date=today)
    total_today = today_attendance.count()
    present_today = today_attendance.filter(status='present').count()
    attendance_percentage = round((present_today / total_today * 100) if total_today > 0 else 0, 1)

    # Recent activities
    recent_activities = Attendance.objects.select_related(
        'student', 'teacher', 'subject'
    ).order_by('-created_at')[:10]

    # Top performing students (by attendance)
    top_students = Profile.objects.filter(role='student').annotate(
        total_attendance=Count('user__attendances'),
        present_count=Count('user__attendances', filter=Q(user__attendances__status='present'))
    ).order_by('-present_count')[:5]

    context = {
        'total_students': total_students,
        'total_teachers': total_teachers,
        'total_classes': total_classes,
        'attendance_percentage': attendance_percentage,
        'recent_activities': recent_activities,
        'top_students': top_students,
        'current_date': today,
    }

    return render(request, 'attendance/manage_users.html', context)


@login_required
def manage_users(request):
    """Manage all users"""
    if not user_is_admin(request.user):
        messages.error(request, "Sizda admin huquqlari yo'q!")
        return redirect('login')

    users = Profile.objects.select_related('user', 'student_class').all()

    # Filters
    role_filter = request.GET.get('role', '')
    status_filter = request.GET.get('status', '')
    class_filter = request.GET.get('class', '')
    search = request.GET.get('search', '')

    if role_filter:
        users = users.filter(role=role_filter)
    if status_filter:
        users = users.filter(user__is_active=(status_filter == 'active'))
    if class_filter:
        users = users.filter(student_class__name=class_filter)
    if search:
        users = users.filter(
            Q(user__username__icontains=search) |
            Q(user__first_name__icontains=search) |
            Q(user__last_name__icontains=search)
        )

    classes = Class.objects.all()

    context = {
        'users': users,
        'classes': classes,
        'role_filter': role_filter,
        'status_filter': status_filter,
        'class_filter': class_filter,
        'search': search,
    }

    return render(request, 'attendance/manage_users.html', context)


@login_required
def add_user(request):
    """Add new user"""
    if not user_is_admin(request.user):
        messages.error(request, "Sizda admin huquqlari yo'q!")
        return redirect('login')

    if request.method == 'POST':
        form = UserForm(request.POST)
        if form.is_valid():
            user = User.objects.create_user(
                username=form.cleaned_data['username'],
                email=form.cleaned_data['email'],
                password=form.cleaned_data['password'],
                first_name=form.cleaned_data['first_name'],
                last_name=form.cleaned_data['last_name'],
            )

            profile = Profile.objects.create(
                user=user,
                role=form.cleaned_data['role'],
                phone=form.cleaned_data.get('phone', ''),
                birth_date=form.cleaned_data.get('birth_date'),
            )

            if form.cleaned_data['role'] == 'student' and form.cleaned_data.get('student_class'):
                profile.student_class = form.cleaned_data['student_class']
                profile.save()

            messages.success(request, f"{user.get_full_name()} muvaffaqiyatli qo'shildi!")
            return redirect('manage_users')
    else:
        form = UserForm()

    return render(request, 'attendance/add_user.html', {'form': form})


@login_required
def edit_user(request, user_id):
    """Edit user"""
    if not user_is_admin(request.user):
        messages.error(request, "Sizda admin huquqlari yo'q!")
        return redirect('login')

    user = get_object_or_404(User, id=user_id)
    profile = user.profile

    if request.method == 'POST':
        user.username = request.POST.get('username')
        user.email = request.POST.get('email')
        user.first_name = request.POST.get('first_name')
        user.last_name = request.POST.get('last_name')

        # Update password if provided
        password = request.POST.get('password')
        if password:
            confirm_password = request.POST.get('confirm_password')
            if password == confirm_password:
                user.set_password(password)
            else:
                messages.error(request, "Parollar mos kelmadi!")
                classes = Class.objects.all()
                context = {
                    'edit_user': user,
                    'profile': profile,
                    'classes': classes,
                }
                return render(request, 'attendance/edit_user.html', context)

        # Update user active status
        user.is_active = request.POST.get('is_active') == 'true'
        user.save()

        # Update profile
        profile.phone = request.POST.get('phone', '')
        profile.role = request.POST.get('role')

        if profile.role == 'student':
            class_id = request.POST.get('student_class')
            if class_id:
                profile.student_class = Class.objects.get(id=class_id)
            else:
                profile.student_class = None
        else:
            profile.student_class = None

        profile.save()

        messages.success(request, f"{user.get_full_name()} muvaffaqiyatli yangilandi!")
        return redirect('manage_users')

    # Get statistics for the user
    if profile.role == 'student':
        # Calculate student attendance statistics
        attendance_records = Attendance.objects.filter(student=user)
        total = attendance_records.count()
        present = attendance_records.filter(status='present').count()
        absent = attendance_records.filter(status='absent').count()
        percentage = round((present / total * 100) if total > 0 else 0, 1)

        student_attendance = {
            'present': present,
            'absent': absent,
            'percentage': percentage,
            'total': total
        }
    elif profile.role == 'teacher':
        # Calculate teacher statistics
        classes_count = Class.objects.filter(teacher=user).count()
        subjects_count = Subject.objects.filter(teacher=user).count()
        attendance_count = Attendance.objects.filter(teacher=user).count()

        teacher_stats = {
            'classes': classes_count,
            'subjects': subjects_count,
            'attendance': attendance_count
        }

    classes = Class.objects.all()
    context = {
        'edit_user': user,
        'profile': profile,
        'classes': classes,
    }

    # Add statistics to context
    if profile.role == 'student':
        context['student_attendance'] = student_attendance
    elif profile.role == 'teacher':
        context['teacher_stats'] = teacher_stats

    return render(request, 'attendance/edit_user.html', context)

@login_required
def delete_user(request, user_id):
    """Delete user"""
    if not user_is_admin(request.user):
        messages.error(request, "Sizda admin huquqlari yo'q!")
        return redirect('login')

    user = get_object_or_404(User, id=user_id)
    username = user.username
    user.delete()

    messages.success(request, f"{username} muvaffaqiyatli o'chirildi!")
    return redirect('manage_users')


@login_required
def manage_classes(request):
    """Manage classes"""
    if not user_is_admin(request.user):
        messages.error(request, "Sizda admin huquqlari yo'q!")
        return redirect('login')

    classes = Class.objects.select_related('teacher').annotate(
        student_count=Count('students')
    ).all()

    # Calculate attendance for each class
    today = timezone.now().date()
    for cls in classes:
        total = Attendance.objects.filter(class_obj=cls, date=today).count()
        present = Attendance.objects.filter(class_obj=cls, date=today, status='present').count()
        cls.attendance_rate = round((present / total * 100) if total > 0 else 0, 1)

    context = {
        'classes': classes,
    }

    return render(request, 'attendance/manage_classes.html', context)


@login_required
def add_class(request):
    """Add new class"""
    if not user_is_admin(request.user):
        messages.error(request, "Sizda admin huquqlari yo'q!")
        return redirect('login')

    if request.method == 'POST':
        form = ClassForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Sinf muvaffaqiyatli qo'shildi!")
            return redirect('manage_classes')
    else:
        form = ClassForm()

    return render(request, 'attendance/add_class.html', {'form': form})


@login_required
def edit_class(request, class_id):
    """Edit class"""
    if not user_is_admin(request.user):
        messages.error(request, "Sizda admin huquqlari yo'q!")
        return redirect('login')

    class_obj = get_object_or_404(Class, id=class_id)

    if request.method == 'POST':
        form = ClassForm(request.POST, instance=class_obj)
        if form.is_valid():
            form.save()
            messages.success(request, "Sinf muvaffaqiyatli yangilandi!")
            return redirect('manage_classes')
    else:
        form = ClassForm(instance=class_obj)

    return render(request, 'attendance/edit_class.html', {'form': form, 'class_obj': class_obj})


@login_required
def delete_class(request, class_id):
    """Delete class"""
    if not user_is_admin(request.user):
        messages.error(request, "Sizda admin huquqlari yo'q!")
        return redirect('login')

    class_obj = get_object_or_404(Class, id=class_id)
    class_name = class_obj.name
    class_obj.delete()

    messages.success(request, f"{class_name} muvaffaqiyatli o'chirildi!")
    return redirect('manage_classes')


@login_required
def manage_subjects(request):
    """Manage subjects"""
    if not user_is_admin(request.user):
        messages.error(request, "Sizda admin huquqlari yo'q!")
        return redirect('login')

    subjects = Subject.objects.select_related('teacher').prefetch_related('classes').all()

    context = {
        'subjects': subjects,
    }

    return render(request, 'attendance/manage_subjects.html', context)


@login_required
def add_subject(request):
    """Add new subject"""
    if not user_is_admin(request.user):
        messages.error(request, "Sizda admin huquqlari yo'q!")
        return redirect('login')

    if request.method == 'POST':
        form = SubjectForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Fan muvaffaqiyatli qo'shildi!")
            return redirect('manage_subjects')
    else:
        form = SubjectForm()

    return render(request, 'attendance/add_subject.html', {'form': form})


@login_required
def edit_subject(request, subject_id):
    """Edit subject"""
    if not user_is_admin(request.user):
        messages.error(request, "Sizda admin huquqlari yo'q!")
        return redirect('login')

    subject = get_object_or_404(Subject, id=subject_id)

    if request.method == 'POST':
        form = SubjectForm(request.POST, instance=subject)
        if form.is_valid():
            form.save()
            messages.success(request, "Fan muvaffaqiyatli yangilandi!")
            return redirect('manage_subjects')
    else:
        form = SubjectForm(instance=subject)

    return render(request, 'attendance/edit_subject.html', {'form': form, 'subject': subject})


@login_required
def delete_subject(request, subject_id):
    """Delete subject"""
    if not user_is_admin(request.user):
        messages.error(request, "Sizda admin huquqlari yo'q!")
        return redirect('login')

    subject = get_object_or_404(Subject, id=subject_id)
    subject_name = subject.name
    subject.delete()

    messages.success(request, f"{subject_name} muvaffaqiyatli o'chirildi!")
    return redirect('manage_subjects')


@login_required
def manage_attendance(request):
    """Manage attendance records"""
    if not user_is_admin(request.user):
        messages.error(request, "Sizda admin huquqlari yo'q!")
        return redirect('login')

    today = timezone.now().date()

    # Get filter parameters
    date_filter = request.GET.get('date', str(today))
    class_filter = request.GET.get('class', '')
    subject_filter = request.GET.get('subject', '')

    try:
        filter_date = datetime.strptime(date_filter, '%Y-%m-%d').date()
    except:
        filter_date = today

    attendance_records = Attendance.objects.filter(date=filter_date).select_related(
        'student', 'subject', 'class_obj', 'teacher'
    )

    if class_filter:
        attendance_records = attendance_records.filter(class_obj__name=class_filter)
    if subject_filter:
        attendance_records = attendance_records.filter(subject__name=subject_filter)

    # Calculate statistics
    present_count = attendance_records.filter(status='present').count()
    absent_count = attendance_records.filter(status='absent').count()
    late_count = attendance_records.filter(status='late').count()
    excused_count = attendance_records.filter(status='excused').count()

    classes = Class.objects.all()
    subjects = Subject.objects.all()

    context = {
        'attendance_records': attendance_records,
        'classes': classes,
        'subjects': subjects,
        'filter_date': filter_date,
        'class_filter': class_filter,
        'subject_filter': subject_filter,
        'present_count': present_count,
        'absent_count': absent_count,
        'late_count': late_count,
        'excused_count': excused_count,
    }

    return render(request, 'attendance/manage_attendance.html', context)


@login_required
def take_attendance(request):
    """Take attendance for a class"""
    if not user_is_admin(request.user):
        messages.error(request, "Sizda admin huquqlari yo'q!")
        return redirect('login')

    if request.method == 'POST':
        class_id = request.POST.get('class_id')
        subject_id = request.POST.get('subject_id')
        date = request.POST.get('date', timezone.now().date())

        if not class_id or not subject_id:
            messages.error(request, "Iltimos, sinf va fanni tanlang!")
            return redirect('take_attendance')

        class_obj = get_object_or_404(Class, id=class_id)
        subject = get_object_or_404(Subject, id=subject_id)

        # Sinfdagi barcha o'quvchilarni olish
        students = Profile.objects.filter(role='student', student_class=class_obj)

        # Davomat yozuvlarini saqlash yoki yangilash
        for student_profile in students:
            status = request.POST.get(f'status_{student_profile.user.id}', 'absent')
            notes = request.POST.get(f'notes_{student_profile.user.id}', '')

            # Davomatni saqlash yoki yangilash
            Attendance.objects.update_or_create(
                student=student_profile.user,
                subject=subject,
                class_obj=class_obj,
                date=date,
                defaults={
                    'teacher': request.user,
                    'status': status,
                    'notes': notes,
                }
            )

        messages.success(request, f"Davomat muvaffaqiyatli saqlandi! {students.count()} ta o'quvchi qayd etildi.")
        return redirect('manage_attendance')

    # GET so'rovi uchun sinf va fanlar ro'yxatini tayyorlash
    classes = Class.objects.all()
    subjects = Subject.objects.all()

    context = {
        'classes': classes,
        'subjects': subjects,
    }

    return render(request, 'attendance/take_attendance.html', context)


@login_required
def attendance_details(request, attendance_id):
    """Get attendance details for AJAX request"""
    if not user_is_admin(request.user):
        return JsonResponse({'error': 'Unauthorized'}, status=403)

    attendance = get_object_or_404(
        Attendance.objects.select_related('student', 'teacher', 'subject', 'class_obj'),
        id=attendance_id
    )

    data = {
        'id': attendance.id,
        'student_name': attendance.student.get_full_name(),
        'student_username': attendance.student.username,
        'class_name': attendance.class_obj.name,
        'subject_name': attendance.subject.name,
        'teacher_name': attendance.teacher.get_full_name() if attendance.teacher else 'Tayinlanmagan',
        'date': attendance.date.strftime("%d.%m.%Y"),
        'status': attendance.status,
        'status_display': attendance.get_status_display(),
        'notes': attendance.notes,
        'created_at': attendance.created_at.strftime("%d.%m.%Y %H:%M"),
        'updated_at': attendance.updated_at.strftime("%d.%m.%Y %H:%M"),
    }

    return JsonResponse(data)


@login_required
def edit_attendance(request, attendance_id):
    """Edit attendance record"""
    if not user_is_admin(request.user):
        messages.error(request, "Sizda admin huquqlari yo'q!")
        return redirect('login')

    attendance = get_object_or_404(Attendance, id=attendance_id)

    if request.method == 'POST':
        attendance.status = request.POST.get('status')
        attendance.notes = request.POST.get('notes', '')
        attendance.save()

        messages.success(request, "Davomat muvaffaqiyatli yangilandi!")
        return redirect('manage_attendance')

    return render(request, 'attendance/edit_attendance.html', {'attendance': attendance})


@login_required
def delete_attendance(request, attendance_id):
    """Delete attendance record"""
    if not user_is_admin(request.user):
        messages.error(request, "Sizda admin huquqlari yo'q!")
        return redirect('login')

    attendance = get_object_or_404(Attendance, id=attendance_id)
    attendance.delete()

    messages.success(request, "Davomat muvaffaqiyatli o'chirildi!")
    return redirect('manage_attendance')


@login_required
@login_required
def statistics(request):
    """Statistics page"""
    if not user_is_admin(request.user):
        messages.error(request, "Sizda admin huquqlari yo'q!")
        return redirect('login')

    # Period filter
    period = request.GET.get('period', 'month')
    class_filter = request.GET.get('class', '')

    today = timezone.now().date()

    if period == 'week':
        start_date = today - timedelta(days=7)
        period_display = "1 hafta"
    elif period == 'month':
        start_date = today - timedelta(days=30)
        period_display = "1 oy"
    elif period == 'quarter':
        start_date = today - timedelta(days=90)
        period_display = "3 oy"
    else:  # year
        start_date = today - timedelta(days=365)
        period_display = "1 yil"

    # Get attendance records
    attendance_records = Attendance.objects.filter(date__gte=start_date)

    if class_filter:
        attendance_records = attendance_records.filter(class_obj__name=class_filter)

    # Calculate statistics
    total_records = attendance_records.count()
    present_records = attendance_records.filter(status='present').count()
    overall_percentage = round((present_records / total_records * 100) if total_records > 0 else 0, 1)

    # Subject-wise statistics
    subject_stats = []
    for subject in Subject.objects.all():
        subject_attendance = attendance_records.filter(subject=subject)
        total = subject_attendance.count()
        present = subject_attendance.filter(status='present').count()
        percentage = round((present / total * 100) if total > 0 else 0, 1)
        subject_stats.append({
            'name': subject.name,
            'percentage': percentage,
            'total': total,
            'present': present
        })

    classes = Class.objects.all()

    context = {
        'overall_percentage': overall_percentage,
        'subject_stats': subject_stats,
        'classes': classes,
        'period': period,
        'class_filter': class_filter,
        'period_display': period_display,
    }

    return render(request, 'attendance/statistics.html', context)


@login_required
def class_stats_json(request):
    """Get class statistics for AJAX request"""
    if not user_is_admin(request.user):
        return JsonResponse({'error': 'Unauthorized'}, status=403)

    period = request.GET.get('period', 'month')
    class_filter = request.GET.get('class_filter', '')

    # Calculate date range based on period
    today = timezone.now().date()
    if period == 'week':
        start_date = today - timedelta(days=7)
    elif period == 'month':
        start_date = today - timedelta(days=30)
    elif period == 'quarter':
        start_date = today - timedelta(days=90)
    else:
        start_date = today - timedelta(days=365)

    # Get class statistics
    classes = Class.objects.all()
    data = {
        'labels': [],
        'data': []
    }

    for class_obj in classes:
        attendance_records = Attendance.objects.filter(
            date__gte=start_date,
            class_obj=class_obj
        )

        if class_filter and class_obj.name != class_filter:
            continue

        total = attendance_records.count()
        present = attendance_records.filter(status='present').count()
        percentage = round((present / total * 100) if total > 0 else 0, 1)

        data['labels'].append(class_obj.name)
        data['data'].append(percentage)

    return JsonResponse(data)


@login_required
def trend_data_json(request):
    """Get trend data for AJAX request"""
    if not user_is_admin(request.user):
        return JsonResponse({'error': 'Unauthorized'}, status=403)

    months = int(request.GET.get('months', 6))
    class_filter = request.GET.get('class_filter', '')

    data = {
        'labels': [],
        'data': []
    }

    today = timezone.now().date()
    for i in range(months - 1, -1, -1):
        month_date = today - timedelta(days=30 * i)
        month_start = month_date.replace(day=1)
        if i == 0:
            month_end = today
        else:
            next_month = month_start + timedelta(days=32)
            month_end = next_month.replace(day=1) - timedelta(days=1)

        # Get attendance for this month
        attendance_records = Attendance.objects.filter(
            date__gte=month_start,
            date__lte=month_end
        )

        if class_filter:
            attendance_records = attendance_records.filter(class_obj__name=class_filter)

        total = attendance_records.count()
        present = attendance_records.filter(status='present').count()
        percentage = round((present / total * 100) if total > 0 else 0, 1)

        data['labels'].append(month_start.strftime('%b %Y'))
        data['data'].append(percentage)

    return JsonResponse(data)


@login_required
def top_students_json(request):
    """Get top students for AJAX request"""
    if not user_is_admin(request.user):
        return JsonResponse({'error': 'Unauthorized'}, status=403)

    period = request.GET.get('period', 'month')
    class_filter = request.GET.get('class_filter', '')

    # Calculate date range
    today = timezone.now().date()
    if period == 'week':
        start_date = today - timedelta(days=7)
    elif period == 'month':
        start_date = today - timedelta(days=30)
    elif period == 'quarter':
        start_date = today - timedelta(days=90)
    else:
        start_date = today - timedelta(days=365)

    # Get student statistics
    students = Profile.objects.filter(role='student')
    student_stats = []

    for profile in students:
        if class_filter and profile.student_class and profile.student_class.name != class_filter:
            continue

        attendance_records = Attendance.objects.filter(
            student=profile.user,
            date__gte=start_date
        )

        total = attendance_records.count()
        present = attendance_records.filter(status='present').count()
        percentage = round((present / total * 100) if total > 0 else 0, 1)

        if total > 0:  # Only include students with attendance records
            student_stats.append({
                'name': profile.user.get_full_name(),
                'class_name': profile.student_class.name if profile.student_class else 'Sinf yo\'q',
                'attendance_percentage': percentage,
                'present_count': present,
                'total_count': total
            })

    # Sort by percentage (highest first) and take top 5
    student_stats.sort(key=lambda x: x['attendance_percentage'], reverse=True)
    top_students = student_stats[:5]

    return JsonResponse(top_students, safe=False)


@login_required
def low_attendance_json(request):
    """Get low attendance students for AJAX request"""
    if not user_is_admin(request.user):
        return JsonResponse({'error': 'Unauthorized'}, status=403)

    period = request.GET.get('period', 'month')
    class_filter = request.GET.get('class_filter', '')

    # Calculate date range
    today = timezone.now().date()
    if period == 'week':
        start_date = today - timedelta(days=7)
    elif period == 'month':
        start_date = today - timedelta(days=30)
    elif period == 'quarter':
        start_date = today - timedelta(days=90)
    else:
        start_date = today - timedelta(days=365)

    # Get student statistics
    students = Profile.objects.filter(role='student')
    low_attendance_students = []

    for profile in students:
        if class_filter and profile.student_class and profile.student_class.name != class_filter:
            continue

        attendance_records = Attendance.objects.filter(
            student=profile.user,
            date__gte=start_date
        )

        total = attendance_records.count()
        present = attendance_records.filter(status='present').count()
        absent = attendance_records.filter(status='absent').count()
        percentage = round((present / total * 100) if total > 0 else 0, 1)

        if total > 0 and percentage < 70:  # Less than 70% attendance
            low_attendance_students.append({
                'name': profile.user.get_full_name(),
                'class_name': profile.student_class.name if profile.student_class else 'Sinf yo\'q',
                'attendance_percentage': percentage,
                'present_count': present,
                'absent_count': absent,
                'total_count': total
            })

    # Sort by percentage (lowest first) and take top 5
    low_attendance_students.sort(key=lambda x: x['attendance_percentage'])
    low_students = low_attendance_students[:5]

    return JsonResponse(low_students, safe=False)


@login_required
def subjects_json(request):
    """Get all subjects for AJAX request"""
    if not user_is_admin(request.user):
        return JsonResponse({'error': 'Unauthorized'}, status=403)

    subjects = Subject.objects.all().values('id', 'name')
    return JsonResponse(list(subjects), safe=False)
@login_required
@login_required
def reports(request):
    """Reports page"""
    if not user_is_admin(request.user):
        messages.error(request, "Sizda admin huquqlari yo'q!")
        return redirect('login')

    classes = Class.objects.all()
    subjects = Subject.objects.all()

    context = {
        'classes': classes,
        'subjects': subjects,
    }

    return render(request, 'attendance/reports.html', context)

@login_required
def preview_report(request):
    """Preview report data via AJAX"""
    if not user_is_admin(request.user):
        return JsonResponse({'error': 'Unauthorized'}, status=403)

    if request.method == 'POST':
        # Get filter parameters
        report_type = request.POST.get('report_type', 'daily')
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        class_id = request.POST.get('class')
        subject_id = request.POST.get('subject')
        status_filter = request.POST.get('status')

        # Build query
        attendance_records = Attendance.objects.all()

        if start_date and end_date:
            attendance_records = attendance_records.filter(
                date__gte=start_date,
                date__lte=end_date
            )

        if class_id:
            attendance_records = attendance_records.filter(class_obj_id=class_id)

        if subject_id:
            attendance_records = attendance_records.filter(subject_id=subject_id)

        if status_filter:
            attendance_records = attendance_records.filter(status=status_filter)

        # Get limited records for preview
        records = attendance_records.select_related('student', 'class_obj', 'subject')[:50]

        # Prepare data
        data = {
            'date_range': f"{start_date} - {end_date}" if start_date and end_date else "Barcha davr",
            'total_records': attendance_records.count(),
            'records': [
                {
                    'student_name': record.student.get_full_name(),
                    'class_name': record.class_obj.name,
                    'subject_name': record.subject.name,
                    'date': record.date.strftime("%d.%m.%Y"),
                    'status': record.status,
                }
                for record in records
            ],
            'summary': {
                'present': attendance_records.filter(status='present').count(),
                'absent': attendance_records.filter(status='absent').count(),
                'late': attendance_records.filter(status='late').count(),
                'excused': attendance_records.filter(status='excused').count(),
                'attendance_percentage': round(
                    (attendance_records.filter(status='present').count() / attendance_records.count() * 100)
                    if attendance_records.count() > 0 else 0, 1
                )
            }
        }

        return JsonResponse(data)

    return JsonResponse({'error': 'Invalid request method'}, status=400)
@login_required
def generate_report(request):
    """Generate and download report"""
    if not user_is_admin(request.user):
        messages.error(request, "Sizda admin huquqlari yo'q!")
        return redirect('login')

    if request.method == 'POST':
        # ... existing code for generating CSV ...

        # Add other formats
        export_format = request.POST.get('export_format', 'csv')

        if export_format == 'excel':
            # Generate Excel file using openpyxl or xlwt
            # Return Excel file
            pass
        elif export_format == 'pdf':
            # Generate PDF using reportlab
            # Return PDF file
            pass
        elif export_format == 'print':
            # Render HTML template for printing
            return render(request, 'attendance/print_report.html', context)

    return redirect('reports')

@login_required
def generate_report(request):
    """Generate and download report"""
    if not user_is_admin(request.user):
        messages.error(request, "Sizda admin huquqlari yo'q!")
        return redirect('login')

    if request.method == 'POST':
        report_type = request.POST.get('report_type')
        class_filter = request.POST.get('class', '')
        subject_filter = request.POST.get('subject', '')
        export_format = request.POST.get('export_format', 'csv')

        # Get date range
        today = timezone.now().date()
        if report_type == 'daily':
            start_date = end_date = today
        elif report_type == 'weekly':
            start_date = today - timedelta(days=7)
            end_date = today
        elif report_type == 'monthly':
            start_date = today - timedelta(days=30)
            end_date = today
        else:
            start_date = request.POST.get('start_date')
            end_date = request.POST.get('end_date')

        # Get attendance records
        attendance_records = Attendance.objects.filter(
            date__gte=start_date,
            date__lte=end_date
        ).select_related('student', 'subject', 'class_obj')

        if class_filter:
            attendance_records = attendance_records.filter(class_obj__name=class_filter)
        if subject_filter:
            attendance_records = attendance_records.filter(subject__name=subject_filter)

        # Generate CSV
        if export_format == 'csv':
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="davomat_hisoboti_{today}.csv"'

            writer = csv.writer(response)
            writer.writerow(['O\'quvchi', 'Sinf', 'Fan', 'Sana', 'Holat', 'Izoh'])

            for record in attendance_records:
                writer.writerow([
                    record.student.get_full_name(),
                    record.class_obj.name,
                    record.subject.name,
                    record.date,
                    record.get_status_display(),
                    record.notes or '',
                ])

            return response

    return redirect('reports')


@login_required
def settings_view(request):
    """Settings page"""
    if not user_is_admin(request.user):
        messages.error(request, "Sizda admin huquqlari yo'q!")
        return redirect('login')

    settings_obj, created = Settings.objects.get_or_create(id=1)

    if request.method == 'POST':
        form = SettingsForm(request.POST, instance=settings_obj)
        if form.is_valid():
            form.save()

            # Change password if provided
            current_password = request.POST.get('current_password')
            new_password = request.POST.get('new_password')
            confirm_password = request.POST.get('confirm_password')

            if current_password and new_password and confirm_password:
                if request.user.check_password(current_password):
                    if new_password == confirm_password:
                        request.user.set_password(new_password)
                        request.user.save()
                        messages.success(request, "Parol muvaffaqiyatli o'zgartirildi!")
                    else:
                        messages.error(request, "Yangi parollar mos kelmadi!")
                else:
                    messages.error(request, "Joriy parol noto'g'ri!")

            messages.success(request, "Sozlamalar muvaffaqiyatli saqlandi!")
            return redirect('settings')
    else:
        form = SettingsForm(instance=settings_obj)

    context = {
        'form': form,
        'settings_obj': settings_obj,
    }

    return render(request, 'attendance/settings.html', context)


@login_required
def get_class_students(request, class_id):
    """AJAX endpoint to get students in a class"""
    if not user_is_admin(request.user):
        return JsonResponse({'error': 'Unauthorized'}, status=403)

    class_obj = get_object_or_404(Class, id=class_id)
    students = Profile.objects.filter(
        role='student',
        student_class=class_obj
    ).select_related('user')

    student_list = [{
        'id': s.user.id,
        'name': s.user.get_full_name(),
        'username': s.user.username,
    } for s in students]

    return JsonResponse({'students': student_list})


@login_required
def get_subject_info(request, subject_id):
    """AJAX endpoint to get subject info"""
    if not user_is_admin(request.user):
        return JsonResponse({'error': 'Unauthorized'}, status=403)

    subject = get_object_or_404(Subject, id=subject_id)

    data = {
        'id': subject.id,
        'name': subject.name,
        'teacher': subject.teacher.get_full_name() if subject.teacher else '',
        'lessons_per_week': subject.lessons_per_week,
        'schedule': subject.schedule,
        'classes': [{'id': c.id, 'name': c.name} for c in subject.classes.all()]
    }

    return JsonResponse(data)


@login_required
def student_dashboard(request):
    """Student dashboard view"""
    if not hasattr(request.user, 'profile') or request.user.profile.role != 'student':
        messages.error(request, "Sizda o'quvchi huquqlari yo'q!")
        return redirect('login')

    # Student's attendance records
    my_attendance = Attendance.objects.filter(
        student=request.user
    ).select_related('subject', 'class_obj').order_by('-date')[:10]

    # Calculate attendance percentage
    total = Attendance.objects.filter(student=request.user).count()
    present = Attendance.objects.filter(student=request.user, status='present').count()
    my_percentage = round((present / total * 100) if total > 0 else 0, 1)

    context = {
        'my_attendance': my_attendance,
        'my_percentage': my_percentage,
    }

    return render(request, 'attendance/student_dashboard.html', context)


# attendance/views.py - teacher_dashboard funksiyasini yangilang

@login_required
def teacher_dashboard(request):
    """Teacher dashboard view"""
    if not hasattr(request.user, 'profile') or request.user.profile.role != 'teacher':
        messages.error(request, "Sizda o'qituvchi huquqlari yo'q!")
        return redirect('login')

    # Teacher's subjects
    my_subjects = Subject.objects.filter(teacher=request.user)

    # Teacher's classes - barcha dars beradigan sinflar
    my_classes = Class.objects.filter(
        Q(teacher=request.user) |  # sinf rahbari bo'lgan
        Q(subjects__teacher=request.user)  # dars beradigan
    ).distinct()

    # Recent attendance taken by this teacher
    recent_attendance = Attendance.objects.filter(
        teacher=request.user
    ).select_related('student', 'subject', 'class_obj').order_by('-created_at')[:10]

    # Today's attendance statistics
    today = timezone.now().date()
    today_attendance = Attendance.objects.filter(teacher=request.user, date=today).count()

    # Calculate attendance percentage
    total_attendance = Attendance.objects.filter(teacher=request.user).count()
    present_attendance = Attendance.objects.filter(teacher=request.user, status='present').count()
    attendance_percentage = round((present_attendance / total_attendance * 100) if total_attendance > 0 else 100, 1)

    context = {
        'my_subjects': my_subjects,
        'my_classes': my_classes,
        'recent_attendance': recent_attendance,
        'today_attendance': today_attendance,
        'attendance_percentage': attendance_percentage,
    }

    return render(request, 'attendance/teacher_dashboard.html', context)


@login_required
def teacher_my_subjects(request):
    """Teacher's subjects page"""
    if not hasattr(request.user, 'profile') or request.user.profile.role != 'teacher':
        messages.error(request, "Sizda o'qituvchi huquqlari yo'q!")
        return redirect('login')

    my_subjects = Subject.objects.filter(teacher=request.user).prefetch_related('classes')

    # Calculate statistics
    total_classes = sum(subject.classes.count() for subject in my_subjects)
    total_hours_per_week = sum(subject.lessons_per_week for subject in my_subjects)

    # Get today's lessons count
    today = timezone.now().date()
    today_lessons = 0  # Sizning jadvalingizga qarab hisoblash kerak

    # Get last attendance for each subject
    for subject in my_subjects:
        last_attendance = Attendance.objects.filter(
            subject=subject,
            teacher=request.user
        ).order_by('-date').first()
        subject.last_attendance = last_attendance

        # Count students in all classes for this subject
        total_students = 0
        for class_obj in subject.classes.all():
            total_students += class_obj.students.count()
        subject.total_students = total_students

    context = {
        'my_subjects': my_subjects,
        'total_classes': total_classes,
        'total_hours_per_week': total_hours_per_week,
        'today_lessons': today_lessons,
    }

    return render(request, 'attendance/teacher_my_subjects.html', context)


@login_required
def teacher_subject_details_json(request, subject_id):
    """Get subject details for AJAX request"""
    if not hasattr(request.user, 'profile') or request.user.profile.role != 'teacher':
        return JsonResponse({'error': 'Unauthorized'}, status=403)

    subject = get_object_or_404(Subject, id=subject_id, teacher=request.user)

    # Get classes for this subject
    classes = subject.classes.all()

    # Calculate attendance statistics
    attendance_records = Attendance.objects.filter(subject=subject, teacher=request.user)
    total = attendance_records.count()
    present = attendance_records.filter(status='present').count()
    absent = attendance_records.filter(status='absent').count()
    percentage = round((present / total * 100) if total > 0 else 0, 1)

    data = {
        'id': subject.id,
        'name': subject.name,
        'lessons_per_week': subject.lessons_per_week,
        'schedule': subject.schedule,
        'classes_count': classes.count(),
        'total_students': sum(cls.students.count() for cls in classes),
        'classes': [
            {
                'id': cls.id,
                'name': cls.name,
                'student_count': cls.students.count(),
                'room': cls.room or '',
            }
            for cls in classes
        ],
        'attendance_stats': {
            'total': total,
            'present': present,
            'absent': absent,
            'percentage': percentage,
        }
    }

    return JsonResponse(data)


@login_required
def teacher_subject_classes_json(request, subject_id):
    """Get classes for a subject"""
    if not hasattr(request.user, 'profile') or request.user.profile.role != 'teacher':
        return JsonResponse({'error': 'Unauthorized'}, status=403)

    subject = get_object_or_404(Subject, id=subject_id, teacher=request.user)

    classes = subject.classes.all()

    data = {
        'classes': [
            {
                'id': cls.id,
                'name': cls.name,
                'student_count': cls.students.count(),
            }
            for cls in classes
        ]
    }

    return JsonResponse(data)


# attendance/views.py - Add these functions

@login_required
@login_required
def teacher_my_classes(request):
    """Teacher's classes page"""
    if not hasattr(request.user, 'profile') or request.user.profile.role != 'teacher':
        messages.error(request, "Sizda o'qituvchi huquqlari yo'q!")
        return redirect('login')

    # Get teacher's all classes (rabbi bo'lgan yoki dars beradigan)
    my_classes = Class.objects.filter(
        Q(teacher=request.user) |
        Q(subjects__teacher=request.user)
    ).distinct()

    # Calculate statistics for each class
    for class_obj in my_classes:
        # Count students
        class_obj.student_count = class_obj.students.count()

        # Calculate today's attendance
        today = timezone.now().date()
        today_attendance = Attendance.objects.filter(
            class_obj=class_obj,
            teacher=request.user,
            date=today
        )
        class_obj.today_present = today_attendance.filter(status='present').count()
        class_obj.today_total = today_attendance.count()

        # Calculate overall attendance percentage for this class
        all_attendance = Attendance.objects.filter(class_obj=class_obj, teacher=request.user)
        total_att = all_attendance.count()
        present_att = all_attendance.filter(status='present').count()
        class_obj.attendance_percentage = round((present_att / total_att * 100) if total_att > 0 else 0, 1)

        # Get subjects taught in this class by this teacher
        class_obj.subjects_taught = Subject.objects.filter(
            teacher=request.user,
            classes=class_obj
        )
        class_obj.subjects_count = class_obj.subjects_taught.count()

    # Get statistics for the page
    total_students = sum(cls.student_count for cls in my_classes)
    total_subjects = sum(cls.subjects_count for cls in my_classes)

    context = {
        'my_classes': my_classes,
        'total_students': total_students,
        'total_subjects': total_subjects,
        'total_classes': my_classes.count(),
    }

    return render(request, 'attendance/teacher_my_classes.html', context)


@login_required
def teacher_class_students_json(request, class_id):
    """Get students in a class for teacher"""
    if not hasattr(request.user, 'profile') or request.user.profile.role != 'teacher':
        return JsonResponse({'error': 'Unauthorized'}, status=403)

    # Check if teacher can teach in this class
    class_obj = get_object_or_404(Class, id=class_id)

    # Teacher bu sinfda dars bera oladimi?
    can_teach = Subject.objects.filter(
        teacher=request.user,
        classes=class_obj
    ).exists() or class_obj.teacher == request.user

    if not can_teach:
        return JsonResponse({'error': 'You are not assigned to this class'}, status=403)

    students = Profile.objects.filter(
        role='student',
        student_class=class_obj
    ).select_related('user')

    # Calculate attendance for each student
    student_list = []
    for profile in students:
        student = profile.user

        # Get attendance records for this student in this class by this teacher
        attendance_records = Attendance.objects.filter(
            student=student,
            class_obj=class_obj,
            teacher=request.user
        )

        total = attendance_records.count()
        present = attendance_records.filter(status='present').count()
        percentage = round((present / total * 100) if total > 0 else 0, 1)

        student_list.append({
            'id': student.id,
            'full_name': student.get_full_name(),
            'username': student.username,
            'phone': profile.phone or '-',
            'attendance_percentage': percentage,
            'present_count': present,
            'total_count': total,
        })

    return JsonResponse({'students': student_list})


@login_required
def teacher_class_details_json(request, class_id):
    """Get detailed class information"""
    if not hasattr(request.user, 'profile') or request.user.profile.role != 'teacher':
        return JsonResponse({'error': 'Unauthorized'}, status=403)

    class_obj = get_object_or_404(Class, id=class_id, teacher=request.user)

    # Get all subjects taught in this class by this teacher
    subjects = Subject.objects.filter(
        teacher=request.user,
        classes=class_obj
    )

    # Get recent attendance records for this class
    recent_attendance = Attendance.objects.filter(
        class_obj=class_obj,
        teacher=request.user
    ).select_related('student', 'subject').order_by('-date', '-created_at')[:10]

    # Calculate attendance statistics
    attendance_records = Attendance.objects.filter(class_obj=class_obj, teacher=request.user)
    total_attendance = attendance_records.count()
    present_attendance = attendance_records.filter(status='present').count()
    absent_attendance = attendance_records.filter(status='absent').count()
    late_attendance = attendance_records.filter(status='late').count()

    data = {
        'class_info': {
            'id': class_obj.id,
            'name': class_obj.name,
            'room': class_obj.room or 'Xona yo\'q',
            'schedule': class_obj.schedule or 'Jadval yo\'q',
            'created_at': class_obj.created_at.strftime("%d.%m.%Y"),
            'student_count': class_obj.students.count(),
        },
        'subjects': [
            {
                'id': subject.id,
                'name': subject.name,
                'lessons_per_week': subject.lessons_per_week,
                'schedule': subject.schedule or '-',
            }
            for subject in subjects
        ],
        'recent_attendance': [
            {
                'student_name': att.student.get_full_name(),
                'subject_name': att.subject.name,
                'date': att.date.strftime("%d.%m.%Y"),
                'status': att.get_status_display(),
                'status_class': att.status,
                'time': att.created_at.strftime("%H:%M"),
            }
            for att in recent_attendance
        ],
        'attendance_stats': {
            'total': total_attendance,
            'present': present_attendance,
            'absent': absent_attendance,
            'late': late_attendance,
            'percentage': round((present_attendance / total_attendance * 100) if total_attendance > 0 else 0, 1),
        }
    }

    return JsonResponse(data)


# attendance/views.py - teacher_take_class_attendance funksiyasini quyidagicha o'zgartiring

@login_required
def teacher_take_class_attendance(request, class_id, subject_id=None):
    """Take attendance for a specific class and subject"""
    if not hasattr(request.user, 'profile') or request.user.profile.role != 'teacher':
        messages.error(request, "Sizda o'qituvchi huquqlari yo'q!")
        return redirect('login')

    # Teacher o'qitadigan sinfni olish (faqat o'z sinflari emas)
    class_obj = get_object_or_404(Class, id=class_id)

    # Teacher bu sinfda dars bera oladimi tekshirish
    can_teach = Subject.objects.filter(
        teacher=request.user,
        classes=class_obj
    ).exists()

    if not can_teach and class_obj.teacher != request.user:
        messages.error(request, "Siz bu sinfda dars bermaysiz!")
        return redirect('teacher_my_classes')

    # Agar subject_id berilmagan bo'lsa
    if not subject_id:
        # Teacherning bu sinfda dars beradigan fanlarini olish
        subjects = Subject.objects.filter(
            teacher=request.user,
            classes=class_obj
        )

        if not subjects.exists():
            messages.error(request, "Siz bu sinfda dars bermaysiz!")
            return redirect('teacher_my_classes')

        if request.method == 'POST':
            subject_id = request.POST.get('subject_id')
            if subject_id:
                return redirect('teacher_take_class_attendance',
                                class_id=class_id,
                                subject_id=subject_id)

        return render(request, 'attendance/teacher_select_subject.html', {
            'class_obj': class_obj,
            'subjects': subjects,
        })

    # Agar subject_id berilgan bo'lsa
    subject = get_object_or_404(Subject, id=subject_id, teacher=request.user, classes=class_obj)

    if request.method == 'POST':
        date_str = request.POST.get('date', timezone.now().date())
        try:
            date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except:
            date = timezone.now().date()

        # Sinfdagi barcha o'quvchilarni olish
        students = Profile.objects.filter(role='student', student_class=class_obj)

        # Davomat yozuvlarini saqlash yoki yangilash
        attendance_count = 0
        for profile in students:
            status = request.POST.get(f'status_{profile.user.id}', 'absent')
            notes = request.POST.get(f'notes_{profile.user.id}', '')

            # Davomatni saqlash yoki yangilash
            attendance, created = Attendance.objects.update_or_create(
                student=profile.user,
                subject=subject,
                class_obj=class_obj,
                teacher=request.user,
                date=date,
                defaults={
                    'status': status,
                    'notes': notes,
                }
            )
            if created:
                attendance_count += 1

        messages.success(request,
                         f"Davomat muvaffaqiyatli saqlandi! {attendance_count} ta o'quvchi qayd etildi.")
        return redirect('teacher_my_classes')

    # GET so'rovi uchun
    students = Profile.objects.filter(role='student', student_class=class_obj).select_related('user')
    today = timezone.now().date()

    # Bugungi davomat yozuvlarini olish
    existing_attendance = {}
    for profile in students:
        att = Attendance.objects.filter(
            student=profile.user,
            subject=subject,
            class_obj=class_obj,
            teacher=request.user,
            date=today
        ).first()
        if att:
            existing_attendance[profile.user.id] = att

    return render(request, 'attendance/teacher_take_attendance.html', {
        'class_obj': class_obj,
        'subject': subject,
        'students': students,
        'existing_attendance': existing_attendance,
        'today': today,
    })


@login_required
def teacher_view_class_attendance(request, class_id):
    """View attendance history for a class"""
    if not hasattr(request.user, 'profile') or request.user.profile.role != 'teacher':
        messages.error(request, "Sizda o'qituvchi huquqlari yo'q!")
        return redirect('login')

    class_obj = get_object_or_404(Class, id=class_id, teacher=request.user)

    # Get filter parameters
    date_filter = request.GET.get('date', '')
    subject_filter = request.GET.get('subject', '')
    student_filter = request.GET.get('student', '')

    # Get attendance records
    attendance_records = Attendance.objects.filter(
        class_obj=class_obj,
        teacher=request.user
    ).select_related('student', 'subject').order_by('-date', '-created_at')

    # Apply filters
    if date_filter:
        try:
            filter_date = datetime.strptime(date_filter, '%Y-%m-%d').date()
            attendance_records = attendance_records.filter(date=filter_date)
        except:
            pass

    if subject_filter:
        attendance_records = attendance_records.filter(subject__name__icontains=subject_filter)

    if student_filter:
        attendance_records = attendance_records.filter(
            Q(student__first_name__icontains=student_filter) |
            Q(student__last_name__icontains=student_filter) |
            Q(student__username__icontains=student_filter)
        )

    # Get subjects for filter dropdown
    subjects = Subject.objects.filter(teacher=request.user, classes=class_obj)

    # Get students for filter dropdown
    students = Profile.objects.filter(role='student', student_class=class_obj).select_related('user')

    context = {
        'class_obj': class_obj,
        'attendance_records': attendance_records,
        'subjects': subjects,
        'students': students,
        'date_filter': date_filter,
        'subject_filter': subject_filter,
        'student_filter': student_filter,
    }

    return render(request, 'attendance/teacher_class_attendance.html', context)


# attendance/views.py - Add these functions

@login_required
def teacher_my_attendance(request):
    """Teacher's attendance records page"""
    if not hasattr(request.user, 'profile') or request.user.profile.role != 'teacher':
        messages.error(request, "Sizda o'qituvchi huquqlari yo'q!")
        return redirect('login')

    # Get filter parameters
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    class_filter = request.GET.get('class', '')
    subject_filter = request.GET.get('subject', '')
    student_filter = request.GET.get('student', '')
    status_filter = request.GET.get('status', '')

    # Get teacher's attendance records
    attendance_records = Attendance.objects.filter(
        teacher=request.user
    ).select_related('student', 'subject', 'class_obj').order_by('-date', '-created_at')

    # Apply filters
    if date_from:
        try:
            attendance_records = attendance_records.filter(date__gte=date_from)
        except:
            pass

    if date_to:
        try:
            attendance_records = attendance_records.filter(date__lte=date_to)
        except:
            pass

    if class_filter:
        attendance_records = attendance_records.filter(class_obj__name__icontains=class_filter)

    if subject_filter:
        attendance_records = attendance_records.filter(subject__name__icontains=subject_filter)

    if student_filter:
        attendance_records = attendance_records.filter(
            Q(student__first_name__icontains=student_filter) |
            Q(student__last_name__icontains=student_filter) |
            Q(student__username__icontains=student_filter)
        )

    if status_filter:
        attendance_records = attendance_records.filter(status=status_filter)

    # Calculate statistics
    total_records = attendance_records.count()
    present_count = attendance_records.filter(status='present').count()
    absent_count = attendance_records.filter(status='absent').count()
    late_count = attendance_records.filter(status='late').count()
    excused_count = attendance_records.filter(status='excused').count()

    # Get unique classes and subjects for filter dropdowns
    classes = Class.objects.filter(teacher=request.user).distinct()
    subjects = Subject.objects.filter(teacher=request.user).distinct()

    # Pagination
    page = request.GET.get('page', 1)
    paginator = Paginator(attendance_records, 50)  # 50 records per page

    try:
        attendance_records_page = paginator.page(page)
    except PageNotAnInteger:
        attendance_records_page = paginator.page(1)
    except EmptyPage:
        attendance_records_page = paginator.page(paginator.num_pages)

    context = {
        'attendance_records': attendance_records_page,
        'classes': classes,
        'subjects': subjects,
        'total_records': total_records,
        'present_count': present_count,
        'absent_count': absent_count,
        'late_count': late_count,
        'excused_count': excused_count,
        'date_from': date_from,
        'date_to': date_to,
        'class_filter': class_filter,
        'subject_filter': subject_filter,
        'student_filter': student_filter,
        'status_filter': status_filter,
    }

    return render(request, 'attendance/teacher_my_attendance.html', context)


@login_required
def teacher_edit_attendance(request, attendance_id):
    """Edit attendance record (teacher version)"""
    if not hasattr(request.user, 'profile') or request.user.profile.role != 'teacher':
        messages.error(request, "Sizda o'qituvchi huquqlari yo'q!")
        return redirect('login')

    attendance = get_object_or_404(Attendance, id=attendance_id, teacher=request.user)

    if request.method == 'POST':
        attendance.status = request.POST.get('status')
        attendance.notes = request.POST.get('notes', '')
        attendance.save()

        messages.success(request, "Davomat muvaffaqiyatli yangilandi!")
        return redirect('teacher_my_attendance')

    # Get student's other attendance records for context
    student_attendance = Attendance.objects.filter(
        student=attendance.student,
        teacher=request.user
    ).exclude(id=attendance_id).order_by('-date')[:5]

    context = {
        'attendance': attendance,
        'student_attendance': student_attendance,
    }

    return render(request, 'attendance/teacher_edit_attendance.html', context)


@login_required
def teacher_delete_attendance(request, attendance_id):
    """Delete attendance record (teacher version)"""
    if not hasattr(request.user, 'profile') or request.user.profile.role != 'teacher':
        messages.error(request, "Sizda o'qituvchi huquqlari yo'q!")
        return redirect('login')

    attendance = get_object_or_404(Attendance, id=attendance_id, teacher=request.user)

    if request.method == 'POST':
        student_name = attendance.student.get_full_name()
        attendance.delete()
        messages.success(request, f"{student_name} uchun davomat yozuvi o'chirildi!")
        return redirect('teacher_my_attendance')

    return render(request, 'attendance/teacher_delete_attendance.html', {'attendance': attendance})


@login_required
def teacher_attendance_details_json(request, attendance_id):
    """Get attendance details for AJAX request"""
    if not hasattr(request.user, 'profile') or request.user.profile.role != 'teacher':
        return JsonResponse({'error': 'Unauthorized'}, status=403)

    attendance = get_object_or_404(
        Attendance.objects.select_related('student', 'subject', 'class_obj'),
        id=attendance_id,
        teacher=request.user
    )

    # Calculate student's overall attendance percentage
    student_attendance = Attendance.objects.filter(
        student=attendance.student,
        teacher=request.user
    )
    total = student_attendance.count()
    present = student_attendance.filter(status='present').count()
    percentage = round((present / total * 100) if total > 0 else 0, 1)

    data = {
        'id': attendance.id,
        'student_name': attendance.student.get_full_name(),
        'student_username': attendance.student.username,
        'student_id': attendance.student.id,
        'class_name': attendance.class_obj.name,
        'class_room': attendance.class_obj.room or '-',
        'subject_name': attendance.subject.name,
        'teacher_name': attendance.teacher.get_full_name(),
        'date': attendance.date.strftime("%d.%m.%Y"),
        'status': attendance.status,
        'status_display': attendance.get_status_display(),
        'notes': attendance.notes or '',
        'created_at': attendance.created_at.strftime("%d.%m.%Y %H:%M"),
        'updated_at': attendance.updated_at.strftime("%d.%m.%Y %H:%M"),
        'student_stats': {
            'total': total,
            'present': present,
            'percentage': percentage,
        }
    }

    return JsonResponse(data)


@login_required
def teacher_export_attendance(request):
    """Export attendance records to CSV/Excel"""
    if not hasattr(request.user, 'profile') or request.user.profile.role != 'teacher':
        messages.error(request, "Sizda o'qituvchi huquqlari yo'q!")
        return redirect('login')

    # Get filter parameters (same as teacher_my_attendance)
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    class_filter = request.GET.get('class', '')
    subject_filter = request.GET.get('subject', '')
    student_filter = request.GET.get('student', '')
    status_filter = request.GET.get('status', '')

    # Get attendance records
    attendance_records = Attendance.objects.filter(
        teacher=request.user
    ).select_related('student', 'subject', 'class_obj').order_by('-date', '-created_at')

    # Apply filters
    if date_from:
        attendance_records = attendance_records.filter(date__gte=date_from)
    if date_to:
        attendance_records = attendance_records.filter(date__lte=date_to)
    if class_filter:
        attendance_records = attendance_records.filter(class_obj__name__icontains=class_filter)
    if subject_filter:
        attendance_records = attendance_records.filter(subject__name__icontains=subject_filter)
    if student_filter:
        attendance_records = attendance_records.filter(
            Q(student__first_name__icontains=student_filter) |
            Q(student__last_name__icontains=student_filter)
        )
    if status_filter:
        attendance_records = attendance_records.filter(status=status_filter)

    # Get export format
    export_format = request.GET.get('format', 'csv')

    if export_format == 'csv':
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="davomat_{timezone.now().date()}.csv"'

        # Write BOM for UTF-8
        response.write('\ufeff')

        writer = csv.writer(response)

        # Write header
        writer.writerow([
            'Sana', 'O\'quvchi', 'Sinf', 'Fan', 'Holat', 'Izoh',
            'Yaratilgan', 'Yangilangan'
        ])

        # Write data
        for record in attendance_records:
            writer.writerow([
                record.date.strftime("%d.%m.%Y"),
                record.student.get_full_name(),
                record.class_obj.name,
                record.subject.name,
                record.get_status_display(),
                record.notes or '',
                record.created_at.strftime("%d.%m.%Y %H:%M"),
                record.updated_at.strftime("%d.%m.%Y %H:%M"),
            ])

        return response

    elif export_format == 'excel':
        # Create Excel file using openpyxl
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="davomat_{timezone.now().date()}.xlsx"'

        wb = Workbook()
        ws = wb.active
        ws.title = "Davomat"

        # Write header
        headers = ['Sana', 'O\'quvchi', 'Sinf', 'Fan', 'Holat', 'Izoh', 'Yaratilgan', 'Yangilangan']
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws[f'{col_letter}1'] = header
            ws[f'{col_letter}1'].font = openpyxl.styles.Font(bold=True)

        # Write data
        for row_num, record in enumerate(attendance_records, 2):
            ws[f'A{row_num}'] = record.date.strftime("%d.%m.%Y")
            ws[f'B{row_num}'] = record.student.get_full_name()
            ws[f'C{row_num}'] = record.class_obj.name
            ws[f'D{row_num}'] = record.subject.name
            ws[f'E{row_num}'] = record.get_status_display()
            ws[f'F{row_num}'] = record.notes or ''
            ws[f'G{row_num}'] = record.created_at.strftime("%d.%m.%Y %H:%M")
            ws[f'H{row_num}'] = record.updated_at.strftime("%d.%m.%Y %H:%M")

            # Add color based on status
            if record.status == 'present':
                ws[f'E{row_num}'].font = openpyxl.styles.Font(color="006400")  # Dark green
            elif record.status == 'absent':
                ws[f'E{row_num}'].font = openpyxl.styles.Font(color="8B0000")  # Dark red
            elif record.status == 'late':
                ws[f'E{row_num}'].font = openpyxl.styles.Font(color="FF8C00")  # Dark orange

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(response)
        return response

    messages.error(request, "Noto'g'ri export format!")
    return redirect('teacher_my_attendance')


@login_required
def teacher_bulk_delete_attendance(request):
    """Bulk delete attendance records"""
    if not hasattr(request.user, 'profile') or request.user.profile.role != 'teacher':
        messages.error(request, "Sizda o'qituvchi huquqlari yo'q!")
        return redirect('login')

    if request.method == 'POST':
        attendance_ids = request.POST.getlist('attendance_ids')

        if not attendance_ids:
            messages.error(request, "Hech qanday davomat yozuvi tanlanmadi!")
            return redirect('teacher_my_attendance')

        # Delete selected attendance records
        deleted_count = Attendance.objects.filter(
            id__in=attendance_ids,
            teacher=request.user
        ).delete()[0]

        messages.success(request, f"{deleted_count} ta davomat yozuvi muvaffaqiyatli o'chirildi!")

    return redirect('teacher_my_attendance')


@login_required
def teacher_today_attendance(request):
    """Get today's attendance records for teacher"""
    if not hasattr(request.user, 'profile') or request.user.profile.role != 'teacher':
        messages.error(request, "Sizda o'qituvchi huquqlari yo'q!")
        return redirect('login')

    today = timezone.now().date()

    # Get today's attendance
    attendance_records = Attendance.objects.filter(
        teacher=request.user,
        date=today
    ).select_related('student', 'subject', 'class_obj').order_by('-created_at')

    # Calculate statistics
    present_count = attendance_records.filter(status='present').count()
    absent_count = attendance_records.filter(status='absent').count()
    late_count = attendance_records.filter(status='late').count()
    total_count = attendance_records.count()

    # Get classes with attendance today
    classes_today = Class.objects.filter(
        teacher=request.user,
        attendance__date=today
    ).distinct().annotate(
        present_count=Count('attendance', filter=Q(attendance__status='present', attendance__date=today)),
        total_count=Count('attendance', filter=Q(attendance__date=today))
    )

    context = {
        'attendance_records': attendance_records,
        'today': today,
        'present_count': present_count,
        'absent_count': absent_count,
        'late_count': late_count,
        'total_count': total_count,
        'classes_today': classes_today,
    }

    return render(request, 'attendance/teacher_today_attendance.html', context)


# attendance/views.py (yoki fayl boshida)
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.contrib.auth.decorators import login_required

# import modelaringizni to'g'ri yo'ldan oling:
from .models import Class, Subject  # agar modellaringiz boshqacha papkada bo'lsa, moslashtiring

@login_required
def teacher_take_attendance_form(request):
    """Show form for taking attendance (for modal)"""
    # profile mavjudligini tekshirish
    if not hasattr(request.user, 'profile') or getattr(request.user.profile, 'role', None) != 'teacher':
        return JsonResponse({'error': 'Unauthorized'}, status=403)

    # Get teacher's classes and subjects
    teacher_classes = Class.objects.filter(teacher=request.user)
    teacher_subjects = Subject.objects.filter(teacher=request.user)

    # Create HTML form
    html = f"""
    <div class="take-attendance-form">
        <h6>Davomat olish uchun sinf va fanni tanlang:</h6>

        <div class="mb-3">
            <label for="class_select" class="form-label">Sinf</label>
            <select class="form-select" id="class_select" onchange="loadSubjectsForClass(this.value)">
                <option value="">Sinfni tanlang</option>
                {"".join(f'<option value="{cls.id}">{cls.name}</option>' for cls in teacher_classes)}
            </select>
        </div>

        <div class="mb-3" id="subject_container" style="display: none;">
            <label for="subject_select" class="form-label">Fan</label>
            <select class="form-select" id="subject_select">
                <option value="">Fanni tanlang</option>
            </select>
        </div>

        <div class="mb-3" id="date_container" style="display: none;">
            <label for="attendance_date" class="form-label">Sana</label>
            <input type="date" class="form-control" id="attendance_date" 
                   value="{timezone.now().date().strftime('%Y-%m-%d')}">
        </div>

        <div class="text-end mt-4">
            <button type="button" class="btn btn-secondary" data-bs-dismiss="modal">Bekor qilish</button>
            <button type="button" class="btn btn-primary" id="start_attendance_btn" 
                    onclick="startAttendance()" disabled>
                <i class="fas fa-clipboard-check me-1"></i>Davomat olish
            </button>
        </div>
    </div>

    <script>
    function loadSubjectsForClass(classId) {{
        if (!classId) {{
            document.getElementById('subject_container').style.display = 'none';
            document.getElementById('date_container').style.display = 'none';
            document.getElementById('start_attendance_btn').disabled = true;
            return;
        }}

        fetch(`/teacher/class/${{classId}}/subjects/`)
            .then(response => response.json())
            .then(data => {{
                const subjectSelect = document.getElementById('subject_select');
                subjectSelect.innerHTML = '<option value="">Fanni tanlang</option>';

                if (data.subjects && data.subjects.length > 0) {{
                    data.subjects.forEach(subject => {{
                        subjectSelect.innerHTML += `<option value="${{subject.id}}">${{subject.name}}</option>`;
                    }});
                    document.getElementById('subject_container').style.display = 'block';
                    document.getElementById('date_container').style.display = 'block';
                }} else {{
                    document.getElementById('subject_container').style.display = 'none';
                    document.getElementById('date_container').style.display = 'none';
                }}

                updateStartButton();
            }})
            .catch(err => {{
                console.error('Subjects load error:', err);
            }});
    }}

    function updateStartButton() {{
        const classId = document.getElementById('class_select').value;
        const subjectId = document.getElementById('subject_select').value;
        const date = document.getElementById('attendance_date').value;

        document.getElementById('start_attendance_btn').disabled = !(classId && subjectId && date);
    }}

    function startAttendance() {{
        const classId = document.getElementById('class_select').value;
        const subjectId = document.getElementById('subject_select').value;

        if (classId && subjectId) {{
            window.location.href = `/teacher/class/${{classId}}/take-attendance/${{subjectId}}/`;
        }}
    }}

    document.addEventListener('DOMContentLoaded', function() {{
        const cls = document.getElementById('class_select');
        const subj = document.getElementById('subject_select');
        const date = document.getElementById('attendance_date');
        if (cls) cls.addEventListener('change', updateStartButton);
        if (subj) subj.addEventListener('change', updateStartButton);
        if (date) date.addEventListener('change', updateStartButton);
    }});
    </script>
    """

    return HttpResponse(html)


# attendance/views.py - Add this function

@login_required
def teacher_class_subjects_json(request, class_id):
    """Get subjects for a class (for teacher)"""
    if not hasattr(request.user, 'profile') or request.user.profile.role != 'teacher':
        return JsonResponse({'error': 'Unauthorized'}, status=403)

    # Check if teacher is assigned to this class
    class_obj = get_object_or_404(Class, id=class_id, teacher=request.user)

    # Get subjects taught in this class by this teacher
    subjects = Subject.objects.filter(
        teacher=request.user,
        classes=class_obj
    )

    data = {
        'subjects': [
            {
                'id': subject.id,
                'name': subject.name,
                'lessons_per_week': subject.lessons_per_week,
                'schedule': subject.schedule or '',
            }
            for subject in subjects
        ]
    }

    return JsonResponse(data)


# attendance/views.py - Add this function

@login_required
def teacher_today_schedule_json(request):
    """Get today's schedule for AJAX request"""
    if not hasattr(request.user, 'profile') or request.user.profile.role != 'teacher':
        return JsonResponse({'error': 'Unauthorized'}, status=403)

    today = timezone.now().date()

    # Get today's attendance records
    attendance_records = Attendance.objects.filter(
        teacher=request.user,
        date=today
    ).select_related('subject', 'class_obj').order_by('created_at')

    # Format data for frontend
    data = {
        'attendance_records': [
            {
                'time': record.created_at.strftime("%H:%M"),
                'subject_name': record.subject.name,
                'class_name': record.class_obj.name,
                'room': record.class_obj.room or '',
                'class_id': record.class_obj.id,
                'subject_id': record.subject.id,
            }
            for record in attendance_records
        ]
    }

    return JsonResponse(data)


# attendance/views.py - Add these imports at the top
import json
from collections import defaultdict
from datetime import datetime, timedelta
from django.db.models import Count, Avg, Q, F
import plotly.graph_objects as go
import plotly.offline as opy
import pandas as pd


# attendance/views.py - Add these functions

@login_required
def teacher_reports(request):
    """Teacher reports dashboard"""
    if not hasattr(request.user, 'profile') or request.user.profile.role != 'teacher':
        messages.error(request, "Sizda o'qituvchi huquqlari yo'q!")
        return redirect('login')

    # Get teacher's subjects and classes
    subjects = Subject.objects.filter(teacher=request.user)
    classes = Class.objects.filter(teacher=request.user)

    # Default date range (last 30 days)
    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=30)

    # Get attendance statistics
    attendance_stats = get_teacher_attendance_stats(request.user, start_date, end_date)

    # Get student performance data
    student_stats = get_teacher_student_stats(request.user, start_date, end_date)

    # Get subject-wise statistics
    subject_stats = get_teacher_subject_stats(request.user, start_date, end_date)

    # Get class-wise statistics
    class_stats = get_teacher_class_stats(request.user, start_date, end_date)

    context = {
        'subjects': subjects,
        'classes': classes,
        'attendance_stats': attendance_stats,
        'student_stats': student_stats,
        'subject_stats': subject_stats,
        'class_stats': class_stats,
        'start_date': start_date,
        'end_date': end_date,
        'today': end_date,
    }

    return render(request, 'attendance/teacher_reports.html', context)


def get_teacher_attendance_stats(teacher, start_date, end_date):
    """Get teacher's overall attendance statistics"""
    attendance_records = Attendance.objects.filter(
        teacher=teacher,
        date__gte=start_date,
        date__lte=end_date
    )

    total = attendance_records.count()
    present = attendance_records.filter(status='present').count()
    absent = attendance_records.filter(status='absent').count()
    late = attendance_records.filter(status='late').count()
    excused = attendance_records.filter(status='excused').count()

    return {
        'total': total,
        'present': present,
        'absent': absent,
        'late': late,
        'excused': excused,
        'percentage': round((present / total * 100) if total > 0 else 0, 1),
        'avg_daily': round(total / max((end_date - start_date).days, 1), 1),
    }


def get_teacher_student_stats(teacher, start_date, end_date):
    """Get student performance statistics"""
    # Get all students taught by this teacher
    students = User.objects.filter(
        attendances__teacher=teacher,
        attendances__date__gte=start_date,
        attendances__date__lte=end_date
    ).distinct()

    student_stats = []
    for student in students[:10]:  # Limit to top 10 for dashboard
        attendance_records = Attendance.objects.filter(
            teacher=teacher,
            student=student,
            date__gte=start_date,
            date__lte=end_date
        )

        total = attendance_records.count()
        present = attendance_records.filter(status='present').count()
        percentage = round((present / total * 100) if total > 0 else 0, 1)

        student_stats.append({
            'id': student.id,
            'name': student.get_full_name(),
            'total': total,
            'present': present,
            'percentage': percentage,
            'class_name': student.profile.student_class.name if hasattr(student,
                                                                        'profile') and student.profile.student_class else 'N/A',
        })

    # Sort by percentage (highest first)
    student_stats.sort(key=lambda x: x['percentage'], reverse=True)

    return student_stats


def get_teacher_subject_stats(teacher, start_date, end_date):
    """Get subject-wise statistics"""
    subjects = Subject.objects.filter(teacher=teacher)

    subject_stats = []
    for subject in subjects:
        attendance_records = Attendance.objects.filter(
            teacher=teacher,
            subject=subject,
            date__gte=start_date,
            date__lte=end_date
        )

        total = attendance_records.count()
        present = attendance_records.filter(status='present').count()
        percentage = round((present / total * 100) if total > 0 else 0, 1)

        # Get classes for this subject
        classes = subject.classes.all()

        subject_stats.append({
            'id': subject.id,
            'name': subject.name,
            'total': total,
            'present': present,
            'percentage': percentage,
            'classes_count': classes.count(),
            'students_count': sum(cls.students.count() for cls in classes),
        })

    return subject_stats


def get_teacher_class_stats(teacher, start_date, end_date):
    """Get class-wise statistics"""
    classes = Class.objects.filter(teacher=teacher)

    class_stats = []
    for class_obj in classes:
        attendance_records = Attendance.objects.filter(
            teacher=teacher,
            class_obj=class_obj,
            date__gte=start_date,
            date__lte=end_date
        )

        total = attendance_records.count()
        present = attendance_records.filter(status='present').count()
        percentage = round((present / total * 100) if total > 0 else 0, 1)

        class_stats.append({
            'id': class_obj.id,
            'name': class_obj.name,
            'total': total,
            'present': present,
            'percentage': percentage,
            'student_count': class_obj.students.count(),
            'room': class_obj.room or 'N/A',
        })

    return class_stats


@login_required
def teacher_generate_report(request):
    """Generate custom report based on filters"""
    if not hasattr(request.user, 'profile') or request.user.profile.role != 'teacher':
        return JsonResponse({'error': 'Unauthorized'}, status=403)

    if request.method == 'POST':
        # Get filter parameters
        report_type = request.POST.get('report_type', 'daily')
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        class_id = request.POST.get('class')
        subject_id = request.POST.get('subject')
        student_id = request.POST.get('student')
        group_by = request.POST.get('group_by', 'day')

        # Validate dates
        if not start_date or not end_date:
            return JsonResponse({'error': 'Iltimos, sanalarni kiriting'}, status=400)

        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        except:
            return JsonResponse({'error': 'Noto\'g\'ri sana formati'}, status=400)

        # Build query
        attendance_records = Attendance.objects.filter(
            teacher=request.user,
            date__gte=start_date,
            date__lte=end_date
        ).select_related('student', 'subject', 'class_obj')

        # Apply filters
        if class_id and class_id != 'all':
            attendance_records = attendance_records.filter(class_obj_id=class_id)

        if subject_id and subject_id != 'all':
            attendance_records = attendance_records.filter(subject_id=subject_id)

        if student_id and student_id != 'all':
            attendance_records = attendance_records.filter(student_id=student_id)

        # Generate report data based on type
        if report_type == 'summary':
            report_data = generate_summary_report(attendance_records, group_by)
        elif report_type == 'detailed':
            report_data = generate_detailed_report(attendance_records)
        elif report_type == 'student':
            report_data = generate_student_report(attendance_records, group_by)
        elif report_type == 'subject':
            report_data = generate_subject_report(attendance_records, group_by)
        else:
            report_data = generate_daily_report(attendance_records)

        # Add metadata
        report_data['metadata'] = {
            'report_type': report_type,
            'start_date': start_date.strftime('%d.%m.%Y'),
            'end_date': end_date.strftime('%d.%m.%Y'),
            'total_records': attendance_records.count(),
            'generated_at': timezone.now().strftime('%d.%m.%Y %H:%M'),
            'teacher_name': request.user.get_full_name(),
        }

        return JsonResponse(report_data)

    return JsonResponse({'error': 'Invalid request method'}, status=400)


def generate_summary_report(attendance_records, group_by='day'):
    """Generate summary report"""
    # Group by date, subject, or class based on parameter
    if group_by == 'day':
        # Group by date
        dates = attendance_records.dates('date', 'day')
        summary = []
        for date in dates:
            day_records = attendance_records.filter(date=date)
            total = day_records.count()
            present = day_records.filter(status='present').count()

            summary.append({
                'period': date.strftime('%d.%m.%Y'),
                'total': total,
                'present': present,
                'absent': day_records.filter(status='absent').count(),
                'percentage': round((present / total * 100) if total > 0 else 0, 1),
            })

    elif group_by == 'subject':
        # Group by subject
        subjects = Subject.objects.filter(
            id__in=attendance_records.values_list('subject', flat=True).distinct()
        )
        summary = []
        for subject in subjects:
            subject_records = attendance_records.filter(subject=subject)
            total = subject_records.count()
            present = subject_records.filter(status='present').count()

            summary.append({
                'period': subject.name,
                'total': total,
                'present': present,
                'absent': subject_records.filter(status='absent').count(),
                'percentage': round((present / total * 100) if total > 0 else 0, 1),
            })

    else:  # group_by == 'class'
        # Group by class
        classes = Class.objects.filter(
            id__in=attendance_records.values_list('class_obj', flat=True).distinct()
        )
        summary = []
        for class_obj in classes:
            class_records = attendance_records.filter(class_obj=class_obj)
            total = class_records.count()
            present = class_records.filter(status='present').count()

            summary.append({
                'period': class_obj.name,
                'total': total,
                'present': present,
                'absent': class_records.filter(status='absent').count(),
                'percentage': round((present / total * 100) if total > 0 else 0, 1),
            })

    # Calculate totals
    totals = {
        'total': sum(item['total'] for item in summary),
        'present': sum(item['present'] for item in summary),
        'absent': sum(item['absent'] for item in summary),
    }

    return {
        'summary': summary,
        'totals': totals,
        'chart_data': prepare_chart_data(summary, group_by),
    }


def generate_detailed_report(attendance_records):
    """Generate detailed attendance report"""
    records = attendance_records.order_by('-date', '-created_at')[:100]  # Limit to 100 records

    detailed_data = []
    for record in records:
        detailed_data.append({
            'date': record.date.strftime('%d.%m.%Y'),
            'student_name': record.student.get_full_name(),
            'class_name': record.class_obj.name,
            'subject_name': record.subject.name,
            'status': record.get_status_display(),
            'status_code': record.status,
            'notes': record.notes or '',
            'time': record.created_at.strftime('%H:%M'),
        })

    return {
        'detailed': detailed_data,
        'count': len(detailed_data),
    }


def generate_student_report(attendance_records, group_by='student'):
    """Generate student performance report"""
    # Get unique students
    student_ids = attendance_records.values_list('student', flat=True).distinct()
    students = User.objects.filter(id__in=student_ids)

    student_data = []
    for student in students:
        student_records = attendance_records.filter(student=student)
        total = student_records.count()
        present = student_records.filter(status='present').count()

        # Get attendance by status
        status_counts = {}
        for status_code, status_name in Attendance.STATUS_CHOICES:
            status_counts[status_code] = student_records.filter(status=status_code).count()

        student_data.append({
            'student_id': student.id,
            'student_name': student.get_full_name(),
            'class_name': student.profile.student_class.name if hasattr(student,
                                                                        'profile') and student.profile.student_class else 'N/A',
            'total': total,
            'present': present,
            'absent': status_counts['absent'],
            'late': status_counts['late'],
            'excused': status_counts['excused'],
            'percentage': round((present / total * 100) if total > 0 else 0, 1),
            'attendance_trend': get_student_attendance_trend(student, attendance_records),
        })

    # Sort by percentage (lowest first to identify problem students)
    student_data.sort(key=lambda x: x['percentage'])

    return {
        'students': student_data,
        'average_percentage': round(
            sum(s['percentage'] for s in student_data) / len(student_data) if student_data else 0, 1),
        'total_students': len(student_data),
    }


def generate_subject_report(attendance_records, group_by='subject'):
    """Generate subject performance report"""
    # Get unique subjects
    subject_ids = attendance_records.values_list('subject', flat=True).distinct()
    subjects = Subject.objects.filter(id__in=subject_ids)

    subject_data = []
    for subject in subjects:
        subject_records = attendance_records.filter(subject=subject)
        total = subject_records.count()
        present = subject_records.filter(status='present').count()

        # Get attendance by class
        class_breakdown = []
        classes = subject.classes.all()
        for class_obj in classes:
            class_records = subject_records.filter(class_obj=class_obj)
            class_total = class_records.count()
            class_present = class_records.filter(status='present').count()

            if class_total > 0:
                class_breakdown.append({
                    'class_name': class_obj.name,
                    'total': class_total,
                    'present': class_present,
                    'percentage': round((class_present / class_total * 100), 1),
                })

        subject_data.append({
            'subject_id': subject.id,
            'subject_name': subject.name,
            'total': total,
            'present': present,
            'percentage': round((present / total * 100) if total > 0 else 0, 1),
            'classes_count': classes.count(),
            'class_breakdown': class_breakdown,
        })

    # Sort by percentage (lowest first)
    subject_data.sort(key=lambda x: x['percentage'])

    return {
        'subjects': subject_data,
        'average_percentage': round(
            sum(s['percentage'] for s in subject_data) / len(subject_data) if subject_data else 0, 1),
        'total_subjects': len(subject_data),
    }


def generate_daily_report(attendance_records):
    """Generate daily attendance report"""
    # Group by date
    dates = attendance_records.dates('date', 'day').order_by('-date')

    daily_data = []
    for date in dates:
        day_records = attendance_records.filter(date=date)
        total = day_records.count()
        present = day_records.filter(status='present').count()

        # Get breakdown by subject
        subject_breakdown = []
        for record in day_records:
            subject_name = record.subject.name
            # Find or create subject entry
            existing = next((s for s in subject_breakdown if s['subject'] == subject_name), None)
            if existing:
                existing['total'] += 1
                if record.status == 'present':
                    existing['present'] += 1
            else:
                subject_breakdown.append({
                    'subject': subject_name,
                    'total': 1,
                    'present': 1 if record.status == 'present' else 0,
                })

        daily_data.append({
            'date': date.strftime('%d.%m.%Y'),
            'day': date.strftime('%A'),
            'total': total,
            'present': present,
            'percentage': round((present / total * 100) if total > 0 else 0, 1),
            'subject_breakdown': subject_breakdown,
        })

    return {
        'daily': daily_data,
        'average_daily': round(sum(d['percentage'] for d in daily_data) / len(daily_data) if daily_data else 0, 1),
        'days_count': len(daily_data),
    }


def get_student_attendance_trend(student, attendance_records):
    """Get student's attendance trend over time"""
    # Get last 7 days
    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=6)

    trend_data = []
    for i in range(7):
        current_date = start_date + timedelta(days=i)
        day_records = attendance_records.filter(
            student=student,
            date=current_date
        )

        total = day_records.count()
        present = day_records.filter(status='present').count()

        trend_data.append({
            'date': current_date.strftime('%d.%m'),
            'percentage': round((present / total * 100) if total > 0 else 0, 0),
            'has_data': total > 0,
        })

    return trend_data


def prepare_chart_data(data, group_by):
    """Prepare data for charts"""
    if not data:
        return {}

    labels = [item['period'] for item in data]
    percentages = [item['percentage'] for item in data]

    # Create Plotly figure
    fig = go.Figure()

    fig.add_trace(go.Bar(
        x=labels,
        y=percentages,
        name='Davomat foizi (%)',
        marker_color='rgb(55, 83, 109)'
    ))

    fig.update_layout(
        title=f'{group_by.capitalize()} bo\'yicha davomat statistikasi',
        xaxis_title=group_by.capitalize(),
        yaxis_title='Davomat foizi (%)',
        height=400,
        showlegend=False
    )

    # Convert to HTML
    chart_html = opy.plot(fig, auto_open=False, output_type='div')

    return {
        'labels': labels,
        'percentages': percentages,
        'chart_html': chart_html,
    }


@login_required
def teacher_export_report(request):
    """Export report in various formats"""
    if not hasattr(request.user, 'profile') or request.user.profile.role != 'teacher':
        messages.error(request, "Sizda o'qituvchi huquqlari yo'q!")
        return redirect('login')

    # Get filter parameters
    report_type = request.GET.get('report_type', 'summary')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    class_id = request.GET.get('class')
    subject_id = request.GET.get('subject')
    format_type = request.GET.get('format', 'pdf')

    # Validate dates
    if not start_date or not end_date:
        messages.error(request, "Iltimos, sanalarni kiriting!")
        return redirect('teacher_reports')

    try:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    except:
        messages.error(request, "Noto'g'ri sana formati!")
        return redirect('teacher_reports')

    # Get attendance records
    attendance_records = Attendance.objects.filter(
        teacher=request.user,
        date__gte=start_date,
        date__lte=end_date
    ).select_related('student', 'subject', 'class_obj')

    # Apply filters
    if class_id and class_id != 'all':
        attendance_records = attendance_records.filter(class_obj_id=class_id)

    if subject_id and subject_id != 'all':
        attendance_records = attendance_records.filter(subject_id=subject_id)

    # Generate report based on type
    if report_type == 'summary':
        report_data = generate_summary_report(attendance_records, 'day')
        filename = f"davomat_hisoboti_{start_date}_{end_date}"
    elif report_type == 'detailed':
        report_data = generate_detailed_report(attendance_records)
        filename = f"batafsil_davomat_{start_date}_{end_date}"
    elif report_type == 'student':
        report_data = generate_student_report(attendance_records)
        filename = f"oquvchi_statistikasi_{start_date}_{end_date}"
    elif report_type == 'subject':
        report_data = generate_subject_report(attendance_records)
        filename = f"fan_statistikasi_{start_date}_{end_date}"
    else:
        report_data = generate_daily_report(attendance_records)
        filename = f"kunlik_davomat_{start_date}_{end_date}"

    # Export based on format
    if format_type == 'csv':
        return export_to_csv(report_data, report_type, filename, request.user)
    elif format_type == 'excel':
        return export_to_excel(report_data, report_type, filename, request.user)
    elif format_type == 'pdf':
        return export_to_pdf(report_data, report_type, filename, request.user)
    else:
        messages.error(request, "Noto'g'ri format turi!")
        return redirect('teacher_reports')


def export_to_csv(report_data, report_type, filename, teacher):
    """Export report to CSV format"""
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'

    # Write BOM for UTF-8
    response.write('\ufeff')
    writer = csv.writer(response)

    if report_type == 'summary':
        writer.writerow(['Davr', 'Jami', 'Qatnashgan', 'Qatnashmagan', 'Foiz (%)'])
        for item in report_data.get('summary', []):
            writer.writerow([
                item['period'],
                item['total'],
                item['present'],
                item['absent'],
                item['percentage']
            ])

    elif report_type == 'detailed':
        writer.writerow(['Sana', 'O\'quvchi', 'Sinf', 'Fan', 'Holat', 'Izoh', 'Vaqt'])
        for item in report_data.get('detailed', []):
            writer.writerow([
                item['date'],
                item['student_name'],
                item['class_name'],
                item['subject_name'],
                item['status'],
                item['notes'],
                item['time']
            ])

    elif report_type == 'student':
        writer.writerow(
            ['O\'quvchi', 'Sinf', 'Jami', 'Qatnashgan', 'Qatnashmagan', 'Kech qolgan', 'Sababli', 'Foiz (%)'])
        for item in report_data.get('students', []):
            writer.writerow([
                item['student_name'],
                item['class_name'],
                item['total'],
                item['present'],
                item['absent'],
                item['late'],
                item['excused'],
                item['percentage']
            ])

    elif report_type == 'subject':
        writer.writerow(['Fan', 'Jami', 'Qatnashgan', 'Foiz (%)', 'Sinflar soni'])
        for item in report_data.get('subjects', []):
            writer.writerow([
                item['subject_name'],
                item['total'],
                item['present'],
                item['percentage'],
                item['classes_count']
            ])

    else:  # daily
        writer.writerow(['Sana', 'Kun', 'Jami', 'Qatnashgan', 'Foiz (%)'])
        for item in report_data.get('daily', []):
            writer.writerow([
                item['date'],
                item['day'],
                item['total'],
                item['present'],
                item['percentage']
            ])

    # Add footer with metadata
    writer.writerow([])
    writer.writerow(['Hisobot ma\'lumotlari:'])
    writer.writerow(['O\'qituvchi:', teacher.get_full_name()])
    writer.writerow(['Yaratilgan:', timezone.now().strftime('%d.%m.%Y %H:%M')])

    return response


def export_to_excel(report_data, report_type, filename, teacher):
    """Export report to Excel format"""
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hisobot"

    # Write header based on report type
    if report_type == 'summary':
        headers = ['Davr', 'Jami', 'Qatnashgan', 'Qatnashmagan', 'Foiz (%)']
        data_key = 'summary'
    elif report_type == 'detailed':
        headers = ['Sana', 'O\'quvchi', 'Sinf', 'Fan', 'Holat', 'Izoh', 'Vaqt']
        data_key = 'detailed'
    elif report_type == 'student':
        headers = ['O\'quvchi', 'Sinf', 'Jami', 'Qatnashgan', 'Qatnashmagan', 'Kech qolgan', 'Sababli', 'Foiz (%)']
        data_key = 'students'
    elif report_type == 'subject':
        headers = ['Fan', 'Jami', 'Qatnashgan', 'Foiz (%)', 'Sinflar soni']
        data_key = 'subjects'
    else:  # daily
        headers = ['Sana', 'Kun', 'Jami', 'Qatnashgan', 'Foiz (%)']
        data_key = 'daily'

    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    # Write data
    data = report_data.get(data_key, [])
    for row_num, item in enumerate(data, 2):
        if report_type == 'summary':
            ws.cell(row=row_num, column=1, value=item['period'])
            ws.cell(row=row_num, column=2, value=item['total'])
            ws.cell(row=row_num, column=3, value=item['present'])
            ws.cell(row=row_num, column=4, value=item['absent'])
            ws.cell(row=row_num, column=5, value=item['percentage'])

        elif report_type == 'detailed':
            ws.cell(row=row_num, column=1, value=item['date'])
            ws.cell(row=row_num, column=2, value=item['student_name'])
            ws.cell(row=row_num, column=3, value=item['class_name'])
            ws.cell(row=row_num, column=4, value=item['subject_name'])
            ws.cell(row=row_num, column=5, value=item['status'])
            ws.cell(row=row_num, column=6, value=item['notes'])
            ws.cell(row=row_num, column=7, value=item['time'])

        elif report_type == 'student':
            ws.cell(row=row_num, column=1, value=item['student_name'])
            ws.cell(row=row_num, column=2, value=item['class_name'])
            ws.cell(row=row_num, column=3, value=item['total'])
            ws.cell(row=row_num, column=4, value=item['present'])
            ws.cell(row=row_num, column=5, value=item['absent'])
            ws.cell(row=row_num, column=6, value=item['late'])
            ws.cell(row=row_num, column=7, value=item['excused'])
            ws.cell(row=row_num, column=8, value=item['percentage'])

            # Color code based on percentage
            if item['percentage'] < 70:
                ws.cell(row=row_num, column=8).font = openpyxl.styles.Font(color="FF0000")
            elif item['percentage'] < 85:
                ws.cell(row=row_num, column=8).font = openpyxl.styles.Font(color="FFA500")
            else:
                ws.cell(row=row_num, column=8).font = openpyxl.styles.Font(color="006400")

        elif report_type == 'subject':
            ws.cell(row=row_num, column=1, value=item['subject_name'])
            ws.cell(row=row_num, column=2, value=item['total'])
            ws.cell(row=row_num, column=3, value=item['present'])
            ws.cell(row=row_num, column=4, value=item['percentage'])
            ws.cell(row=row_num, column=5, value=item['classes_count'])

        else:  # daily
            ws.cell(row=row_num, column=1, value=item['date'])
            ws.cell(row=row_num, column=2, value=item['day'])
            ws.cell(row=row_num, column=3, value=item['total'])
            ws.cell(row=row_num, column=4, value=item['present'])
            ws.cell(row=row_num, column=5, value=item['percentage'])

    # Add metadata
    metadata_row = len(data) + 4
    ws.cell(row=metadata_row, column=1, value="Hisobot ma'lumotlari:").font = openpyxl.styles.Font(bold=True)
    ws.cell(row=metadata_row + 1, column=1, value="O'qituvchi:")
    ws.cell(row=metadata_row + 1, column=2, value=teacher.get_full_name())
    ws.cell(row=metadata_row + 2, column=1, value="Yaratilgan:")
    ws.cell(row=metadata_row + 2, column=2, value=timezone.now().strftime('%d.%m.%Y %H:%M'))

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(response)
    return response


def export_to_pdf(report_data, report_type, filename, teacher):
    """Export report to PDF format"""
    # For PDF generation, we'll use a simple HTML to PDF approach
    # In production, you might want to use reportlab or weasyprint
    # For now, we'll create an HTML response that can be printed

    context = {
        'report_data': report_data,
        'report_type': report_type,
        'teacher': teacher,
        'generated_at': timezone.now().strftime('%d.%m.%Y %H:%M'),
        'filename': filename,
    }

    return render(request, 'attendance/teacher_report_pdf.html', context)


@login_required
def teacher_report_chart_data(request):
    """Get chart data for AJAX requests"""
    if not hasattr(request.user, 'profile') or request.user.profile.role != 'teacher':
        return JsonResponse({'error': 'Unauthorized'}, status=403)

    chart_type = request.GET.get('type', 'overall')
    period = request.GET.get('period', '30')  # days

    # Calculate date range
    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=int(period))

    if chart_type == 'overall':
        data = get_overall_chart_data(request.user, start_date, end_date)
    elif chart_type == 'subject':
        data = get_subject_chart_data(request.user, start_date, end_date)
    elif chart_type == 'class':
        data = get_class_chart_data(request.user, start_date, end_date)
    elif chart_type == 'trend':
        data = get_trend_chart_data(request.user, start_date, end_date)
    else:
        data = {'error': 'Invalid chart type'}

    return JsonResponse(data)


def get_overall_chart_data(teacher, start_date, end_date):
    """Get overall attendance chart data"""
    attendance_records = Attendance.objects.filter(
        teacher=teacher,
        date__gte=start_date,
        date__lte=end_date
    )

    # Status distribution
    status_counts = {
        'present': attendance_records.filter(status='present').count(),
        'absent': attendance_records.filter(status='absent').count(),
        'late': attendance_records.filter(status='late').count(),
        'excused': attendance_records.filter(status='excused').count(),
    }

    # Daily trend
    dates = []
    percentages = []
    current_date = start_date
    while current_date <= end_date:
        day_records = attendance_records.filter(date=current_date)
        total = day_records.count()
        present = day_records.filter(status='present').count()

        dates.append(current_date.strftime('%d.%m'))
        percentages.append(round((present / total * 100) if total > 0 else 0, 1))

        current_date += timedelta(days=1)

    return {
        'status_counts': status_counts,
        'dates': dates,
        'percentages': percentages,
    }


def get_subject_chart_data(teacher, start_date, end_date):
    """Get subject-wise chart data"""
    subjects = Subject.objects.filter(teacher=teacher)

    subject_names = []
    subject_percentages = []

    for subject in subjects:
        attendance_records = Attendance.objects.filter(
            teacher=teacher,
            subject=subject,
            date__gte=start_date,
            date__lte=end_date
        )

        total = attendance_records.count()
        present = attendance_records.filter(status='present').count()
        percentage = round((present / total * 100) if total > 0 else 0, 1)

        subject_names.append(subject.name)
        subject_percentages.append(percentage)

    return {
        'subject_names': subject_names,
        'subject_percentages': subject_percentages,
    }


def get_class_chart_data(teacher, start_date, end_date):
    """Get class-wise chart data"""
    classes = Class.objects.filter(teacher=teacher)

    class_names = []
    class_percentages = []

    for class_obj in classes:
        attendance_records = Attendance.objects.filter(
            teacher=teacher,
            class_obj=class_obj,
            date__gte=start_date,
            date__lte=end_date
        )

        total = attendance_records.count()
        present = attendance_records.filter(status='present').count()
        percentage = round((present / total * 100) if total > 0 else 0, 1)

        class_names.append(class_obj.name)
        class_percentages.append(percentage)

    return {
        'class_names': class_names,
        'class_percentages': class_percentages,
    }


def get_trend_chart_data(teacher, start_date, end_date):
    """Get attendance trend data"""
    # Group by week
    import math

    weeks = []
    week_percentages = []

    current_date = start_date
    week_num = 1

    while current_date <= end_date:
        week_end = min(current_date + timedelta(days=6), end_date)

        week_records = Attendance.objects.filter(
            teacher=teacher,
            date__gte=current_date,
            date__lte=week_end
        )

        total = week_records.count()
        present = week_records.filter(status='present').count()
        percentage = round((present / total * 100) if total > 0 else 0, 1)

        weeks.append(f"Hafta {week_num}")
        week_percentages.append(percentage)

        current_date = week_end + timedelta(days=1)
        week_num += 1

    return {
        'weeks': weeks,
        'week_percentages': week_percentages,
    }


# attendance/views.py - qo'shimcha funksiyalar

@login_required
def teacher_take_attendance_form(request):
    """Simple attendance form for modal"""
    if not hasattr(request.user, 'profile') or request.user.profile.role != 'teacher':
        return HttpResponse('Unauthorized', status=403)

    # Teacher's classes
    teacher_classes = Class.objects.filter(
        Q(teacher=request.user) |
        Q(subjects__teacher=request.user)
    ).distinct()

    html = f"""
    <div class="take-attendance-form">
        <p class="text-muted mb-3">Davomat olish uchun sinf va fanni tanlang:</p>

        <div class="mb-3">
            <label class="form-label">Sinf</label>
            <select class="form-select" id="classSelect" onchange="loadSubjectsForClass(this.value)">
                <option value="">Sinfni tanlang</option>
                {"".join(f'<option value="{cls.id}">{cls.name}</option>' for cls in teacher_classes)}
            </select>
        </div>

        <div class="mb-3" id="subjectContainer" style="display: none;">
            <label class="form-label">Fan</label>
            <select class="form-select" id="subjectSelect">
                <option value="">Fanni tanlang</option>
            </select>
        </div>

        <div class="mb-3" id="dateContainer" style="display: none;">
            <label class="form-label">Sana</label>
            <input type="date" class="form-control" id="attendanceDate" 
                   value="{timezone.now().date().strftime('%Y-%m-%d')}">
        </div>

        <div class="d-grid gap-2 d-md-flex justify-content-md-end mt-4">
            <button type="button" class="btn btn-secondary" data-bs-dismiss="modal">Bekor qilish</button>
            <button type="button" class="btn btn-primary" id="startAttendanceBtn" 
                    onclick="startAttendance()" disabled>
                <i class="fas fa-clipboard-check me-1"></i>Davomat olish
            </button>
        </div>
    </div>

    <script>
    document.getElementById('classSelect').addEventListener('change', updateStartButton);
    document.getElementById('subjectSelect').addEventListener('change', updateStartButton);
    document.getElementById('attendanceDate').addEventListener('change', updateStartButton);
    </script>
    """

    return HttpResponse(html)


@login_required
def teacher_subject_classes_json(request, subject_id):
    """Get classes for a subject"""
    if not hasattr(request.user, 'profile') or request.user.profile.role != 'teacher':
        return JsonResponse({'error': 'Unauthorized'}, status=403)

    subject = get_object_or_404(Subject, id=subject_id, teacher=request.user)
    classes = subject.classes.all()

    data = {
        'classes': [
            {
                'id': cls.id,
                'name': cls.name,
                'room': cls.room or '',
                'student_count': cls.students.count(),
            }
            for cls in classes
        ]
    }

    return JsonResponse(data)


@login_required
def teacher_today_schedule_json(request):
    """Get today's schedule"""
    if not hasattr(request.user, 'profile') or request.user.profile.role != 'teacher':
        return JsonResponse({'error': 'Unauthorized'}, status=403)

    today = timezone.now().date()

    # Get today's attendance records
    attendance_records = Attendance.objects.filter(
        teacher=request.user,
        date=today
    ).select_related('subject', 'class_obj').order_by('created_at')

    data = {
        'attendance_records': [
            {
                'time': record.created_at.strftime("%H:%M"),
                'subject_name': record.subject.name,
                'class_name': record.class_obj.name,
                'class_id': record.class_obj.id,
                'subject_id': record.subject.id,
                'room': record.class_obj.room or '',
            }
            for record in attendance_records
        ]
    }

    return JsonResponse(data)


@login_required
def teacher_take_class_attendance(request, class_id, subject_id):
    """Take attendance for a class and subject"""
    if not hasattr(request.user, 'profile') or request.user.profile.role != 'teacher':
        messages.error(request, "Sizda o'qituvchi huquqlari yo'q!")
        return redirect('login')

    class_obj = get_object_or_404(Class, id=class_id)
    subject = get_object_or_404(Subject, id=subject_id, teacher=request.user, classes=class_obj)

    if request.method == 'POST':
        date_str = request.POST.get('date', timezone.now().date())
        try:
            date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except:
            date = timezone.now().date()

        # Get all students in the class
        students = Profile.objects.filter(role='student', student_class=class_obj)

        # Save attendance for each student
        saved_count = 0
        for profile in students:
            status = request.POST.get(f'status_{profile.user.id}', 'absent')
            notes = request.POST.get(f'notes_{profile.user.id}', '')

            # Create or update attendance
            attendance, created = Attendance.objects.update_or_create(
                student=profile.user,
                subject=subject,
                class_obj=class_obj,
                teacher=request.user,
                date=date,
                defaults={
                    'status': status,
                    'notes': notes,
                }
            )
            if created:
                saved_count += 1

        messages.success(request, f"{saved_count} ta o'quvchi uchun davomat saqlandi!")
        return redirect('teacher_dashboard')

    # GET request - show form
    students = Profile.objects.filter(role='student', student_class=class_obj).select_related('user')
    today = timezone.now().date()

    # Get existing attendance for today
    existing_attendance = {}
    for profile in students:
        att = Attendance.objects.filter(
            student=profile.user,
            subject=subject,
            class_obj=class_obj,
            teacher=request.user,
            date=today
        ).first()
        if att:
            existing_attendance[profile.user.id] = att

    return render(request, 'attendance/take_attendance_form.html', {
        'class_obj': class_obj,
        'subject': subject,
        'students': students,
        'existing_attendance': existing_attendance,
        'today': today,
    })


# views.py
def teacher_edit_attendance(request, attendance_id):
    try:
        attendance = Attendance.objects.get(id=attendance_id, teacher=request.user.teacher)
        attendance_records = AttendanceRecord.objects.filter(attendance=attendance)

        # Statistikani hisoblash
        present_count = attendance_records.filter(status='present').count()
        absent_count = attendance_records.filter(status='absent').count()
        late_count = attendance_records.filter(status='late').count()
        excused_count = attendance_records.filter(status='excused').count()

        context = {
            'attendance': attendance,
            'attendance_records': attendance_records,
            'present_count': present_count,
            'absent_count': absent_count,
            'late_count': late_count,
            'excused_count': excused_count,
        }

        if request.method == 'POST':
            # POST ma'lumotlarini qayta ishlash
            # Sana
            date = request.POST.get('date')
            attendance.date = date

            # Har bir o'quvchi uchun davomat holati
            for record in attendance_records:
                status_key = f'status_{record.student.id}'
                comment_key = f'comment_{record.student.id}'

                if status_key in request.POST:
                    record.status = request.POST[status_key]

                if comment_key in request.POST:
                    record.comment = request.POST[comment_key]

                record.save()

            attendance.save()
            messages.success(request, "Davomat muvaffaqiyatli yangilandi!")
            return redirect('teacher_dashboard')  # yoki kerakli sahifaga

        return render(request, 'attendance/teacher_edit_attendance.html', context)

    except Attendance.DoesNotExist:
        messages.error(request, "Davomat topilmadi!")
        return redirect('teacher_dashboard')